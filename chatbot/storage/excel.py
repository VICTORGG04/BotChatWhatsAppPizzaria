import os
import logging
from datetime import datetime
from openpyxl import Workbook, load_workbook
from models import DadosPedido

logger = logging.getLogger(__name__)


def registrar_pedido_excel(dados: DadosPedido) -> bool:
    """Salva pedido em planilha Excel local diária."""
    arquivo_excel = f'pedidos_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
    
    itens_formatados = ", ".join([
        f"{item.quantidade}x {item.sabor} ({item.tamanho.value})" 
        for item in dados.itens
    ])
    
    nova_linha = [
        dados.numero_do_dia,
        dados.timestamp.strftime("%H:%M:%S"),
        itens_formatados,
        dados.total,
        dados.lucro,
        dados.pagamento.value,
        dados.endereco,
        dados.observacoes
    ]
    
    try:
        if not os.path.exists(arquivo_excel):
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Pedidos"
            cabecalho = ["Nº Pedido (Dia)", "Hora", "Itens", "Total (R$)", 
                        "Lucro (R$)", "Pagamento", "Endereço", "Observações"]
            sheet.append(cabecalho)
            _formatar_cabecalho(sheet)
        else:
            workbook = load_workbook(arquivo_excel)
            if "Pedidos" in workbook.sheetnames:
                sheet = workbook["Pedidos"]
            else:
                sheet = workbook.active
        
        sheet.append(nova_linha)
        _formatar_linha_dados(sheet, sheet.max_row)
        workbook.save(arquivo_excel)
        
        logger.info(f"Pedido #{dados.numero_do_dia} salvo no Excel: {arquivo_excel}")
        return True
        
    except PermissionError:
        logger.error(f"Arquivo Excel '{arquivo_excel}' está aberto em outro programa")
        return False
    except Exception as e:
        logger.error(f"Erro ao salvar no Excel: {e}")
        return False


def _formatar_cabecalho(sheet):
    """Formata linha de cabeçalho."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3CB371", end_color="3CB371", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
    
    sheet.column_dimensions['C'].width = 50
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['G'].width = 50
    sheet.column_dimensions['H'].width = 35


def _formatar_linha_dados(sheet, row_num: int):
    """Formata linha de dados."""
    from openpyxl.styles import PatternFill, Border, Side, numbers
    
    light_green = PatternFill(start_color="D9F2E9", end_color="D9F2E9", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in sheet[row_num]:
        cell.fill = light_green
        cell.border = thin_border
    
    sheet.cell(row=row_num, column=4).number_format = 'R$ #,##0.00'
    sheet.cell(row=row_num, column=5).number_format = 'R$ #,##0.00'