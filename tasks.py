import json
import logging
from datetime import datetime
from typing import Dict, Any

from celery import shared_task
from celery.exceptions import MaxRetriesExceededError

from config import settings
from models import DadosPedido, ItemPedido

logger = logging.getLogger(__name__)


@shared_task(
    bind=True,
    autoretry_for=(Exception,),
    retry_backoff=True,
    retry_backoff_max=600,
    retry_kwargs={"max_retries": 3},
    name="tasks.registrar_pedido_sqlite"
)
def registrar_pedido_sqlite(self, pedido_data: Dict[str, Any]) -> bool:
    """
    Registra pedido no SQLite de forma assíncrona.
    """
    import sqlite3
    
    try:
        pedido = DadosPedido(**pedido_data)
        
        conn = sqlite3.connect(settings.sqlite_db_path)
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO pedidos (numero_do_dia, timestamp, itens_pedido, total_pedido, lucro_pedido, 
                               metodo_pagamento, endereco, observacoes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            pedido.numero_do_dia,
            pedido.timestamp.isoformat(),
            json.dumps([item.model_dump() for item in pedido.itens]),
            pedido.total,
            pedido.lucro,
            pedido.pagamento.value,
            pedido.endereco,
            pedido.observacoes
        ))
        conn.commit()
        conn.close()
        
        logger.info(f"Pedido #{pedido.numero_do_dia} salvo no SQLite")
        return True
        
    except Exception as e:
        logger.error(f"Erro ao salvar no SQLite: {e}")
        raise


@shared_task(
    bind=True,
    autoretry_for=(Exception,),
    retry_backoff=True,
    retry_backoff_max=600,
    retry_kwargs={"max_retries": 3},
    name="tasks.registrar_pedido_google_sheets"
)
def registrar_pedido_google_sheets(self, pedido_data: Dict[str, Any]) -> bool:
    """
    Registra pedido no Google Sheets de forma assíncrona.
    """
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        
        pedido = DadosPedido(**pedido_data)
        
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive"
        ]
        creds = Credentials.from_service_account_file(
            settings.google_credentials_path, scopes=scopes
        )
        gc = gspread.authorize(creds)
        spreadsheet = gc.open(settings.google_sheets_name)
        worksheet = spreadsheet.sheet1
        
        all_values = worksheet.get_all_values()
        is_new_sheet = len(all_values) == 0
        
        if is_new_sheet:
            cabecalho = ["Nº Pedido (Dia)", "Data", "Hora", "Itens", "Total (R$)", 
                        "Lucro (R$)", "Pagamento", "Endereço", "Observações"]
            worksheet.append_row(cabecalho)
        
        itens_formatados = ", ".join([
            f"{item.quantidade}x {item.sabor} ({item.tamanho.value})" 
            for item in pedido.itens
        ])
        
        nova_linha = [
            pedido.numero_do_dia,
            pedido.timestamp.strftime("%d/%m/%Y"),
            pedido.timestamp.strftime("%H:%M:%S"),
            itens_formatados,
            pedido.total,
            pedido.lucro,
            pedido.pagamento.value,
            pedido.endereco,
            pedido.observacoes
        ]
        
        worksheet.append_row(nova_linha)
        nova_linha_numero = len(worksheet.get_all_values())
        
        # Formatação básica
        cor_verde_agua = {"red": 0.85, "green": 0.96, "blue": 0.93}
        borda_solida = {"style": "SOLID", "width": 1, "color": {"red": 0.7, "green": 0.7, "blue": 0.7}}
        
        requests = []
        requests.append({
            "repeatCell": {
                "range": {
                    "sheetId": worksheet.id,
                    "startRowIndex": nova_linha_numero - 1,
                    "endRowIndex": nova_linha_numero
                },
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColor": cor_verde_agua,
                        "borders": {
                            "top": borda_solida, "bottom": borda_solida,
                            "left": borda_solida, "right": borda_solida
                        }
                    }
                },
                "fields": "userEnteredFormat(backgroundColor,borders)"
            }
        })
        
        if is_new_sheet:
            requests.append({
                "repeatCell": {
                    "range": {"sheetId": worksheet.id, "startRowIndex": 0, "endRowIndex": 1},
                    "cell": {
                        "userEnteredFormat": {
                            "backgroundColor": {"red": 0.49, "green": 0.75, "blue": 0.65},
                            "textFormat": {"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
                            "borders": {"top": borda_solida, "bottom": borda_solida, 
                                      "left": borda_solida, "right": borda_solida}
                        }
                    },
                    "fields": "userEnteredFormat(backgroundColor,textFormat,borders)"
                }
            })
        
        if requests:
            spreadsheet.batch_update({"requests": requests})
        
        worksheet.format(f'E{nova_linha_numero}:F{nova_linha_numero}', 
                        {"numberFormat": {"type": "CURRENCY", "pattern": "R$ #,##0.00"}})
        
        logger.info(f"Pedido #{pedido.numero_do_dia} salvo no Google Sheets")
        return True
        
    except Exception as e:
        logger.error(f"Erro ao salvar no Google Sheets: {e}")
        raise


@shared_task(
    bind=True,
    autoretry_for=(Exception,),
    retry_backoff=True,
    retry_backoff_max=300,
    retry_kwargs={"max_retries": 2},
    name="tasks.registrar_pedido_excel"
)
def registrar_pedido_excel(self, pedido_data: Dict[str, Any]) -> bool:
    """
    Registra pedido no Excel local de forma assíncrona.
    """
    try:
        from openpyxl import Workbook, load_workbook
        
        pedido = DadosPedido(**pedido_data)
        arquivo_excel = f'pedidos_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
        
        itens_formatados = ", ".join([
            f"{item.quantidade}x {item.sabor} ({item.tamanho.value})" 
            for item in pedido.itens
        ])
        
        nova_linha = [
            pedido.numero_do_dia,
            pedido.timestamp.strftime("%H:%M:%S"),
            itens_formatados,
            pedido.total,
            pedido.lucro,
            pedido.pagamento.value,
            pedido.endereco,
            pedido.observacoes
        ]
        
        try:
            if not __import__('os').path.exists(arquivo_excel):
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Pedidos"
                cabecalho = ["Nº Pedido (Dia)", "Hora", "Itens", "Total (R$)", 
                           "Lucro (R$)", "Pagamento", "Endereço", "Observações"]
                sheet.append(cabecalho)
            else:
                workbook = load_workbook(arquivo_excel)
                sheet = workbook["Pedidos"] if "Pedidos" in workbook.sheetnames else workbook.active
            
            sheet.append(nova_linha)
            workbook.save(arquivo_excel)
            
            logger.info(f"Pedido #{pedido.numero_do_dia} salvo no Excel local")
            return True
            
        except PermissionError:
            logger.error(f"Arquivo Excel '{arquivo_excel}' está aberto em outro programa")
            raise
        except Exception as e:
            logger.error(f"Erro ao salvar no Excel: {e}")
            raise
            
    except Exception as e:
        logger.error(f"Erro na task Excel: {e}")
        raise


@shared_task(
    bind=True,
    name="tasks.obter_proximo_numero_pedido"
)
def obter_proximo_numero_pedido(self) -> int:
    """
    Obtém próximo número de pedido do dia.
    """
    import sqlite3
    
    try:
        conn = sqlite3.connect(settings.sqlite_db_path)
        cursor = conn.cursor()
        cursor.execute(
            "SELECT MAX(numero_do_dia) FROM pedidos WHERE DATE(timestamp) = DATE('now', 'localtime')"
        )
        ultimo_numero = cursor.fetchone()[0]
        conn.close()
        return (ultimo_numero or 0) + 1
    except Exception as e:
        logger.error(f"Erro ao obter número do pedido: {e}")
        return 1


@shared_task(
    bind=True,
    name="tasks.calcular_lucro_total_dia"
)
def calcular_lucro_total_dia(self) -> float:
    """
    Calcula lucro total do dia.
    """
    import sqlite3
    
    try:
        conn = sqlite3.connect(settings.sqlite_db_path)
        cursor = conn.cursor()
        cursor.execute(
            "SELECT SUM(lucro_pedido) FROM pedidos WHERE DATE(timestamp) = DATE('now', 'localtime')"
        )
        lucro_total = cursor.fetchone()[0]
        conn.close()
        return float(lucro_total or 0.0)
    except Exception as e:
        logger.error(f"Erro ao calcular lucro do dia: {e}")
        return 0.0


@shared_task(
    name="tasks.limpar_sessoes_expiradas"
)
def limpar_sessoes_expiradas() -> int:
    """
    Task periódica para limpar sessões expiradas do Redis.
    """
    from session_manager import session_manager
    import redis
    
    if not session_manager.health_check():
        return 0
    
    try:
        # Redis já expira automaticamente via TTL, mas podemos fazer scan
        # para limpar chaves órfãs se necessário
        pattern = f"{session_manager.SESSION_PREFIX}*"
        cursor = 0
        deleted = 0
        
        while True:
            cursor, keys = session_manager._redis.scan(cursor, match=pattern, count=100)
            if keys:
                # Verifica TTL e deleta chaves sem TTL (órfãs)
                for key in keys:
                    ttl = session_manager._redis.ttl(key)
                    if ttl == -1:  # Sem expiração
                        session_manager._redis.delete(key)
                        deleted += 1
            
            if cursor == 0:
                break
        
        if deleted:
            logger.info(f"Limpas {deleted} sessões órfãs do Redis")
        return deleted
        
    except Exception as e:
        logger.error(f"Erro ao limpar sessões: {e}")
        return 0


@shared_task(
    bind=True,
    name="tasks.enviar_notificacao_pedido"
)
def enviar_notificacao_pedido(self, pedido_data: Dict[str, Any], telefone_atendente: str) -> bool:
    """
    Envia notificação WhatsApp para atendente (futuro).
    """
    # TODO: Implementar envio via WhatsApp Business API
    logger.info(f"Notificação de pedido #{pedido_data.get('numero_do_dia')} para {telefone_atendente}")
    return True