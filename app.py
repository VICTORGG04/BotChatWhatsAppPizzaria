import os
import base64
import logging
from datetime import datetime, timedelta
import jwt as pyjwt
from flask import Flask, request, jsonify, abort

from config import settings
from models import SessaoCliente, WebhookPayload
from session_manager import session_manager
from security import whatsapp_webhook_required, verify_whatsapp_webhook_challenge


def _setup_credentials():
    creds_b64 = os.environ.get("GOOGLE_CREDENTIALS_B64", "")
    if creds_b64 and not os.path.exists("credentials.json"):
        try:
            content = base64.b64decode(creds_b64).decode("utf-8")
            with open("credentials.json", "w") as f:
                f.write(content)
        except Exception:
            pass


_setup_credentials()

import structlog

structlog.configure(
    processors=[
        structlog.stdlib.filter_by_level,
        structlog.stdlib.add_logger_name,
        structlog.stdlib.add_log_level,
        structlog.stdlib.PositionalArgumentsFormatter(),
        structlog.processors.TimeStamper(fmt="iso"),
        structlog.processors.StackInfoRenderer(),
        structlog.processors.format_exc_info,
        structlog.processors.UnicodeDecoder(),
        structlog.processors.JSONRenderer()
    ],
    context_class=dict,
    logger_factory=structlog.stdlib.LoggerFactory(),
    wrapper_class=structlog.stdlib.BoundLogger,
    cache_logger_on_first_use=True,
)

logging.basicConfig(
    format="%(message)s",
    stream=None,
    level=getattr(logging, settings.log_level.upper(), logging.INFO),
)

logger = structlog.get_logger(__name__)

app = Flask(__name__)

with app.app_context():
    try:
        from chatbot.storage.sqlite import setup_database
        setup_database()
    except Exception:
        pass

# ─── Landing ────────────────────────────────────────────────
INDEX_HTML = """<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>PayPizzas - Peça sua pizza pelo WhatsApp!</title>
<style>
@keyframes grad{0%{background-position:0% 50%}50%{background-position:100% 50%}100%{background-position:0% 50%}}
@keyframes float{0%,100%{transform:translateY(0)}50%{transform:translateY(-8px)}}
@keyframes pulse{0%,100%{opacity:1}50%{opacity:.6}}
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;min-height:100vh;display:flex;align-items:center;justify-content:center;background:linear-gradient(135deg,#667eea,#764ba2,#e74c3c,#c0392b);background-size:400% 400%;animation:grad 8s ease infinite}
.card{background:rgba(255,255,255,.95);backdrop-filter:blur(10px);border-radius:24px;padding:48px 40px;text-align:center;max-width:420px;width:90%;box-shadow:0 20px 60px rgba(0,0,0,.25);transition:transform .3s}
.card:hover{transform:scale(1.02)}
.logo{font-size:52px;margin-bottom:8px;display:block}
h1{font-size:28px;color:#e74c3c;margin-bottom:8px;letter-spacing:-.5px}
p{color:#666;line-height:1.6;font-size:15px;margin-bottom:24px}
.status{display:inline-flex;align-items:center;gap:6px;padding:6px 16px;border-radius:20px;font-size:13px;font-weight:500;margin-bottom:8px;transition:all .3s}
.status.aberta{background:#e8f8ef;color:#27ae60}
.status.fechada{background:#fde8e8;color:#e74c3c}
.status .dot{width:8px;height:8px;border-radius:50%;animation:pulse 1.5s ease-in-out infinite}
.status.aberta .dot{background:#27ae60}
.status.fechada .dot{background:#e74c3c}
.horario{font-size:12px;color:#bbb;margin-bottom:20px}
a{display:inline-flex;align-items:center;gap:8px;padding:14px 32px;background:linear-gradient(135deg,#e74c3c,#c0392b);color:#fff;border-radius:14px;text-decoration:none;font-weight:600;font-size:16px;transition:all .25s;box-shadow:0 4px 16px rgba(231,76,60,.35)}
a:hover{transform:translateY(-2px);box-shadow:0 8px 30px rgba(231,76,60,.45)}
a:active{transform:translateY(0);opacity:.7}
a.disabled{opacity:.4;cursor:not-allowed;transform:none!important;box-shadow:none!important}
small{display:block;margin-top:20px;color:#bbb;font-size:12px}
@media(max-width:480px){.card{padding:32px 24px}h1{font-size:24px}}
</style>
</head>
<body>
<div class="card">
<span class="logo">🍕</span>
<h1>PayPizzas</h1>
<p>Faça seu pedido pelo WhatsApp de forma inteligente!</p>
<div class="status" id="statusLoja"><span class="dot"></span><span id="statusTexto">Verificando...</span></div>
<div class="horario" id="horarioLoja"></div>
<a href="/cardapio" id="btnPedir">Fazer Pedido →</a>
<small>PayPizzas Bot v2.0.0</small>
</div>
<script>
fetch('/api/loja/status').then(r=>r.json()).then(d=>{
const el=document.getElementById('statusLoja'),tx=document.getElementById('statusTexto'),hr=document.getElementById('horarioLoja'),btn=document.getElementById('btnPedir');
el.className='status '+(d.aberta?'aberta':'fechada');
tx.textContent=d.aberta?'Loja Aberta':'Loja Fechada';
hr.textContent='Horário: '+d.horario;
if(!d.aberta){btn.href='#';btn.classList.add('disabled');btn.onclick=function(e){e.preventDefault();alert('A loja está fechada. Volte durante o horário de funcionamento.')}}
}).catch(()=>{document.getElementById('statusTexto').textContent='Sistema Online';document.getElementById('statusLoja').className='status aberta'})
</script>
</body>
</html>"""


@app.route('/', methods=['GET'])
def index():
    return INDEX_HTML, 200, {'Content-Type': 'text/html'}


@app.route('/health', methods=['GET'])
def health_check():
    redis_ok = session_manager.health_check()
    return jsonify({
        "status": "healthy" if redis_ok else "degraded",
        "service": "pizzaria-bot",
        "version": "2.0.0",
        "redis": "connected" if redis_ok else "disconnected"
    }), 200 if redis_ok else 503


@app.route('/api/loja/status', methods=['GET'])
def loja_status():
    aberta = settings.is_loja_aberta()
    return jsonify({
        "aberta": aberta,
        "horario": f"{settings.loja_abertura} às {settings.loja_fechamento}"
    })

# ─── Auth / User API ────────────────────────────────────────

_JWT_SECRET = os.environ.get("JWT_SECRET", "pizzaria-secret-change-in-production")
_JWT_ALGO = "HS256"


def _gerar_token(user_id: int) -> str:
    return pyjwt.encode({"sub": user_id, "exp": datetime.utcnow() + timedelta(days=30)}, _JWT_SECRET, algorithm=_JWT_ALGO)


def _token_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.headers.get("Authorization", "")
        if not auth.startswith("Bearer "):
            return jsonify({"erro": "Token nao informado"}), 401
        try:
            payload = pyjwt.decode(auth[7:], _JWT_SECRET, algorithms=[_JWT_ALGO])
            request.user_id = payload["sub"]
        except pyjwt.ExpiredSignatureError:
            return jsonify({"erro": "Token expirado"}), 401
        except Exception:
            return jsonify({"erro": "Token invalido"}), 401
        return f(*args, **kwargs)
    return decorated


@app.route('/api/cardapio', methods=['GET'])
def api_cardapio():
    import importlib.util
    import os as _os
    _base = _os.path.dirname(_os.path.abspath(__file__))
    spec = importlib.util.spec_from_file_location("cardapio_mod", _os.path.join(_base, "chatbot", "cardapio.py"))
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    cardapio = mod.cardapio
    data = {
        "sabores": {},
        "bebidas": cardapio.bebidas,
        "adicionais": cardapio.adicionais,
    }
    for nome, sabor in cardapio.sabores.items():
        tamanhos = {}
        for tam, info in sabor.tamanhos.items():
            tamanhos[tam.value] = {"preco": info.preco, "custo": info.custo, "descricao": info.descricao}
        data["sabores"][nome] = {
            "nome": nome,
            "nome_formatado": sabor.nome_formatado,
            "categoria": sabor.categoria,
            "nome_exibicao": sabor.nome_exibicao,
            "tamanhos": tamanhos,
        }
    return jsonify(data)


@app.route('/api/auth/register', methods=['POST'])
def api_auth_register():
    from chatbot.storage.user_db import criar_usuario
    data = request.get_json()
    if not data:
        return jsonify({"erro": "Dados obrigatorios"}), 400
    email = (data.get("email") or "").strip()
    senha = (data.get("senha") or "").strip()
    nome = (data.get("nome") or "").strip()
    if not email or not senha:
        return jsonify({"erro": "Email e senha obrigatorios"}), 400
    if len(senha) < 6:
        return jsonify({"erro": "Senha deve ter no minimo 6 caracteres"}), 400
    try:
        user = criar_usuario(email, senha, nome)
        token = _gerar_token(user.id)
        return jsonify({"token": token, "user": user.model_dump()}), 201
    except ValueError as e:
        return jsonify({"erro": str(e)}), 409


@app.route('/api/auth/login', methods=['POST'])
def api_auth_login():
    from chatbot.storage.user_db import autenticar_usuario
    data = request.get_json()
    if not data:
        return jsonify({"erro": "Dados obrigatorios"}), 400
    email = (data.get("email") or "").strip()
    senha = (data.get("senha") or "").strip()
    if not email or not senha:
        return jsonify({"erro": "Email e senha obrigatorios"}), 400
    try:
        user = autenticar_usuario(email, senha)
        token = _gerar_token(user.id)
        return jsonify({"token": token, "user": user.model_dump()})
    except ValueError as e:
        return jsonify({"erro": str(e)}), 401


@app.route('/api/auth/me', methods=['GET'])
@_token_required
def api_auth_me():
    from chatbot.storage.user_db import buscar_usuario_por_id
    try:
        user = buscar_usuario_por_id(request.user_id)
        return jsonify({"user": user.model_dump()})
    except ValueError as e:
        return jsonify({"erro": str(e)}), 404


@app.route('/api/auth/me', methods=['PUT'])
@_token_required
def api_auth_update():
    from chatbot.storage.user_db import atualizar_usuario
    data = request.get_json() or {}
    nome = data.get("nome")
    telefone = data.get("telefone")
    try:
        user = atualizar_usuario(request.user_id, nome=nome, telefone=telefone)
        return jsonify({"user": user.model_dump()})
    except ValueError as e:
        return jsonify({"erro": str(e)}), 404


@app.route('/api/auth/enderecos', methods=['GET'])
@_token_required
def api_listar_enderecos():
    from chatbot.storage.user_db import listar_enderecos
    enderecos = listar_enderecos(request.user_id)
    return jsonify({"enderecos": [e.model_dump() for e in enderecos]})


@app.route('/api/auth/enderecos', methods=['POST'])
@_token_required
def api_adicionar_endereco():
    from chatbot.storage.user_db import adicionar_endereco
    data = request.get_json()
    if not data or not data.get("endereco"):
        return jsonify({"erro": "Endereco obrigatorio"}), 400
    endereco = adicionar_endereco(request.user_id, data["endereco"].strip())
    return jsonify({"endereco": endereco.model_dump()}), 201


@app.route('/api/auth/enderecos/<int:id>', methods=['DELETE'])
@_token_required
def api_remover_endereco(id):
    from chatbot.storage.user_db import remover_endereco
    remover_endereco(request.user_id, id)
    return jsonify({"ok": True})


@app.route('/api/auth/favoritos', methods=['GET'])
@_token_required
def api_listar_favoritos():
    from chatbot.storage.user_db import listar_favoritos
    favoritos = listar_favoritos(request.user_id)
    return jsonify({"favoritos": [f.model_dump() for f in favoritos]})


@app.route('/api/auth/favoritos', methods=['POST'])
@_token_required
def api_adicionar_favorito():
    from chatbot.storage.user_db import adicionar_favorito
    data = request.get_json()
    if not data or not data.get("tipo") or not data.get("item_key"):
        return jsonify({"erro": "tipo e item_key obrigatorios"}), 400
    favorito = adicionar_favorito(request.user_id, data["tipo"], data["item_key"], data.get("nome", ""))
    return jsonify({"favorito": favorito.model_dump()}), 201


@app.route('/api/auth/favoritos/<int:id>', methods=['DELETE'])
@_token_required
def api_remover_favorito(id):
    from chatbot.storage.user_db import remover_favorito
    remover_favorito(request.user_id, id)
    return jsonify({"ok": True})


# ─── WhatsApp Webhook ───────────────────────────────────────

@app.route('/webhook', methods=['GET'])
def webhook_verify():
    result = verify_whatsapp_webhook_challenge()
    if result:
        challenge, status = result
        return challenge, status
    return "Webhook endpoint", 200


@app.route('/webhook', methods=['POST'])
@whatsapp_webhook_required
def webhook():
    try:
        payload = request.get_json()
        if not payload:
            logger.warning("Payload vazio recebido")
            abort(400, description="Payload invalido")

        logger.info("Webhook recebido", payload_keys=list(payload.keys()))
        webhook_data = WebhookPayload(**payload)

        for message in webhook_data.get_messages():
            phone_number = message.from_
            text = message.body
            if not phone_number or not text:
                continue

            from chatbot import processar_mensagem
            sessao = session_manager.get(phone_number)
            sessao.chat_history.append({"role": "user", "parts": [{"text": text}]})
            response, updated_sessao = processar_mensagem(text, sessao)
            updated_sessao.chat_history.append({"role": "model", "parts": [{"text": response}]})
            session_manager.save(phone_number, updated_sessao)

            logger.info("Mensagem processada", phone=phone_number[-4:], state=updated_sessao.state, response_length=len(response))

        return "OK", 200

    except Exception as e:
        logger.error("Erro no webhook", error=str(e), exc_info=True)
        abort(500, description="Erro interno do servidor")

# ─── Cardapio Web ───────────────────────────────────────────

@app.route('/cardapio', methods=['GET'])
@app.route('/cardapio/', methods=['GET'])
@app.route('/cardapio/<path:subpath>', methods=['GET'])
def cardapio_page(subpath=None):
    from flask import send_from_directory
    static_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "frontend", "dist")
    if os.path.exists(os.path.join(static_dir, "index.html")):
        if subpath and os.path.exists(os.path.join(static_dir, subpath)):
            return send_from_directory(static_dir, subpath)
        return send_from_directory(static_dir, "index.html")
    # Fallback to server-generated HTML
    import importlib.util
    import os as _os
    _base = _os.path.dirname(_os.path.abspath(__file__))
    spec = importlib.util.spec_from_file_location("cardapio_mod", _os.path.join(_base, "chatbot", "cardapio.py"))
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    from cardapio_html import gerar_cardapio_html
    return gerar_cardapio_html(settings.empresa_numero, mod.cardapio), 200, {'Content-Type': 'text/html; charset=utf-8'}


@app.route('/api/pedido/criar', methods=['POST'])
def api_criar_pedido():
    from datetime import datetime
    from models import DadosPedido, ItemPedido, MetodoPagamento, StatusPedido, TamanhoPizza
    from chatbot.storage.sqlite import registrar_pedido_sqlite, obter_proximo_numero_pedido_dia
    from chatbot.cardapio import cardapio

    try:
        data = request.get_json()
        if not data:
            abort(400, description="Dados invalidos")

        nome = (data.get("nome") or "").strip()
        endereco = (data.get("endereco") or "").strip()
        pagamento_str = (data.get("pagamento") or "").strip()
        observacoes = (data.get("observacoes") or "").strip()
        itens_data = data.get("itens") or []

        if not nome:
            return jsonify({"erro": "Nome é obrigatório"}), 400
        if not endereco:
            return jsonify({"erro": "Endereço é obrigatório"}), 400
        if not pagamento_str:
            return jsonify({"erro": "Forma de pagamento é obrigatória"}), 400
        if not itens_data:
            return jsonify({"erro": "Adicione ao menos um item"}), 400

        pagamento = MetodoPagamento(pagamento_str)
    except Exception:
        abort(400, description="Dados invalidos")

    itens_pedido = []
    total = 0.0
    custo_total = 0.0

    for item in itens_data:
        categoria = item.get("categoria", "sabor").strip().lower()
        sabor_nome = item.get("sabor", "").strip()
        tamanho = item.get("tamanho", "").strip().upper()
        quantidade = int(item.get("quantidade", 1))

        if categoria == "sabor":
            sabor_key = cardapio.buscar_sabor(sabor_nome)
            if not sabor_key:
                continue
            sabor_obj = cardapio.sabores[sabor_key]
            tam_enum = next((t for t in [TamanhoPizza.M, TamanhoPizza.G, TamanhoPizza.GG] if t.value == tamanho), None)
            if not tam_enum or tam_enum not in sabor_obj.tamanhos:
                continue
            item_cardapio = sabor_obj.tamanhos[tam_enum]
        elif categoria == "bebida":
            bebida = next((b for b in cardapio.bebidas if b["chave"] == sabor_nome), None)
            if not bebida:
                continue
            tam_enum = TamanhoPizza.NA
            item_cardapio = type("obj", (object,), {"preco": bebida["preco"], "custo": bebida.get("custo", 0)})()
            sabor_nome = bebida["nome"]
        elif categoria == "adicional":
            adicional = next((a for a in cardapio.adicionais if a["chave"] == sabor_nome), None)
            if not adicional:
                continue
            tam_enum = TamanhoPizza.NA
            item_cardapio = type("obj", (object,), {"preco": adicional["preco"], "custo": adicional.get("custo", 0)})()
            sabor_nome = adicional["nome"]
        else:
            continue

        itens_pedido.append(ItemPedido(
            sabor=sabor_nome,
            tamanho=tam_enum,
            quantidade=quantidade,
            preco=item_cardapio.preco
        ))
        total += item_cardapio.preco * quantidade
        custo_total += item_cardapio.custo * quantidade

    if not itens_pedido:
        return jsonify({"erro": "Nenhum item valido encontrado"}), 400

    numero = obter_proximo_numero_pedido_dia()
    agora = datetime.now()

    dados = DadosPedido(
        numero_do_dia=numero,
        timestamp=agora,
        itens=itens_pedido,
        total=round(total, 2),
        lucro=round(total - custo_total, 2),
        pagamento=pagamento,
        endereco=endereco,
        observacoes=observacoes if observacoes else f"Cliente: {nome}",
        status=StatusPedido.RECEBIDO
    )

    try:
        registrar_pedido_sqlite(dados)
    except Exception as e:
        logger.error("Erro ao salvar pedido no SQLite", error=str(e))
        return jsonify({"erro": "Erro ao salvar pedido"}), 500

    try:
        from tasks import registrar_pedido_google_sheets, registrar_pedido_excel
        pedido_dict = dados.model_dump()
        pedido_dict["timestamp"] = dados.timestamp.isoformat()
        registrar_pedido_google_sheets.delay(pedido_dict)
        registrar_pedido_excel.delay(pedido_dict)
    except Exception:
        pass

    linhas_msg = [f"*NOVO PEDIDO #{numero}*", f"Cliente: {nome}", ""]
    for item in itens_pedido:
        if item.tamanho.value != "NA":
            linhas_msg.append(f"{item.quantidade}x {item.sabor} ({item.tamanho.value}) - R$ {item.preco:.2f}")
        else:
            linhas_msg.append(f"{item.quantidade}x {item.sabor} - R$ {item.preco:.2f}")
    linhas_msg.append("")
    linhas_msg.append(f"*Total: R$ {total:.2f}*")
    linhas_msg.append(f"*Pagamento:* {pagamento.value}")
    linhas_msg.append(f"*Endereco:* {endereco}")
    if observacoes:
        linhas_msg.append(f"*Obs:* {observacoes}")

    texto = "\n".join(linhas_msg)
    import urllib.parse
    wa_link = f"https://wa.me/{settings.empresa_numero}?text={urllib.parse.quote(texto)}"

    return jsonify({
        "numero_do_dia": numero,
        "total": round(total, 2),
        "whatsapp_link": wa_link
    }), 201

# ─── Admin ──────────────────────────────────────────────────

@app.route('/admin', methods=['GET'])
def admin_dashboard():
    from chatbot.admin_html import ADMIN_HTML
    return ADMIN_HTML, 200, {'Content-Type': 'text/html; charset=utf-8'}


@app.route('/admin/pedidos', methods=['GET'])
def admin_pedidos():
    from chatbot.storage.sqlite import buscar_pedidos_hoje
    pedidos = buscar_pedidos_hoje()
    return jsonify({"total": len(pedidos), "pedidos": pedidos})


@app.route('/admin/pedido/<int:numero>/status', methods=['POST'])
def admin_atualizar_status(numero):
    from chatbot.storage.sqlite import atualizar_status_pedido
    data = request.get_json()
    if not data or "status" not in data:
        abort(400, description="Status nao informado")
    status = data["status"].strip()
    valido = atualizar_status_pedido(numero, status)
    if not valido:
        return jsonify({"erro": "Pedido nao encontrado"}), 404
    logger.info("Status atualizado", numero=numero, status=status)
    return jsonify({"ok": True, "numero": numero, "status": status})


@app.route('/admin/stats', methods=['GET'])
def admin_stats():
    from chatbot.storage.sqlite import calcular_lucro_total_dia
    lucro = calcular_lucro_total_dia()
    return jsonify({"lucro_dia": lucro, "moeda": "BRL"})


@app.route('/admin/exportar-excel', methods=['GET'])
def admin_exportar_excel():
    from chatbot.storage.sqlite import buscar_pedidos_hoje
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    import json

    pedidos = buscar_pedidos_hoje()
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos"
    headers = ["N Pedido", "Hora", "Itens", "Total (R$)", "Lucro (R$)", "Pagamento", "Endereco", "Observacoes", "Status"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3CB371", end_color="3CB371", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    light_green = PatternFill(start_color="D9F2E9", end_color="D9F2E9", fill_type="solid")
    for row_idx, pedido in enumerate(pedidos, 2):
        itens = pedido.get("itens_pedido", "[]")
        if isinstance(itens, str):
            try:
                itens_lista = json.loads(itens)
                itens_formatados = ", ".join([f"{i.get('quantidade', '?')}x {i.get('sabor', '?')}" for i in itens_lista])
            except json.JSONDecodeError:
                itens_formatados = itens
        else:
            itens_formatados = str(itens)

        linha = [
            pedido.get("numero_do_dia"),
            pedido.get("timestamp", "")[11:16] if pedido.get("timestamp") else "",
            itens_formatados,
            pedido.get("total_pedido", 0),
            pedido.get("lucro_pedido", 0),
            pedido.get("metodo_pagamento", ""),
            pedido.get("endereco", ""),
            pedido.get("observacoes", ""),
            pedido.get("status", "recebido"),
        ]

        for col, valor in enumerate(linha, 1):
            cell = ws.cell(row=row_idx, column=col, value=valor)
            cell.fill = light_green
            cell.border = thin_border

        ws.cell(row=row_idx, column=4).number_format = 'R$ #,##0.00'
        ws.cell(row=row_idx, column=5).number_format = 'R$ #,##0.00'

    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['G'].width = 50
    ws.column_dimensions['H'].width = 35

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    from flask import send_file
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"pedidos_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    )

# ─── PIX ────────────────────────────────────────────────────

@app.route('/pix/qrcode', methods=['GET'])
def pix_qrcode():
    from flask import send_file
    from chatbot.storage.pix import gerar_pix_qrcode, gerar_pix_payload
    import io

    amount = request.args.get("amount", type=float)
    desc = request.args.get("desc", "")

    if request.args.get("format") == "brcode":
        brcode = gerar_pix_payload(amount=amount, description=desc or None)
        return brcode, 200, {'Content-Type': 'text/plain; charset=utf-8'}

    img_bytes = gerar_pix_qrcode(amount=amount, description=desc or None)
    if img_bytes is None:
        return jsonify({"error": "qrcode nao instalado"}), 500
    return send_file(io.BytesIO(img_bytes), mimetype="image/png")

# ─── Error Handlers ─────────────────────────────────────────

@app.errorhandler(400)
def bad_request(e):
    logger.warning("Bad request", error=str(e))
    return jsonify({"error": "Requisicao invalida", "message": str(e)}), 400


@app.errorhandler(401)
def unauthorized(e):
    logger.warning("Unauthorized", error=str(e))
    return jsonify({"error": "Nao autorizado", "message": str(e)}), 401


@app.errorhandler(500)
def internal_error(e):
    logger.error("Internal server error", error=str(e), exc_info=True)
    return jsonify({"error": "Erro interno", "message": "Erro no processamento"}), 500


if __name__ == '__main__':
    from chatbot.storage.sqlite import setup_database
    setup_database()
    logger.info("Iniciando Pizzaria Bot", host=settings.host, port=settings.port)
    app.run(host=settings.host, port=settings.port, debug=settings.debug)
