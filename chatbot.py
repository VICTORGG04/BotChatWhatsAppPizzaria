# chatbot.py

import os
import sqlite3
import json
from datetime import datetime
from dotenv import load_dotenv

import gspread
from openpyxl import Workbook, load_workbook

import google.generativeai as genai

# --- CONFIGURAÇÃO INICIAL ---
load_dotenv()

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
if GEMINI_API_KEY:
    genai.configure(api_key=GEMINI_API_KEY)
    genai_model = genai.GenerativeModel('gemini-pro')
else:
    print("AVISO: GEMINI_API_KEY não configurada. A funcionalidade de resumo da IA não estará disponível.")
    genai_model = None

# --- DADOS DO SISTEMA ---
CARDAPIO = {
    "margherita": {
        "M": {"preco": 35.00, "custo": 12.50, "descricao": "Molho, muçarela e manjericão."},
        "G": {"preco": 45.00, "custo": 16.00, "descricao": "Molho, muçarela e manjericão."},
        "GG": {"preco": 55.00, "custo": 20.00, "descricao": "Molho, muçarela e manjericão."}
    },
    "calabresa": {
        "M": {"preco": 40.00, "custo": 15.00, "descricao": "Molho, muçarela, calabresa e cebola."},
        "G": {"preco": 52.00, "custo": 19.50, "descricao": "Molho, muçarela, calabresa e cebola."},
        "GG": {"preco": 63.00, "custo": 24.00, "descricao": "Molho, muçarela, calabresa e cebola."}
    },
    "portuguesa": {
        "M": {"preco": 45.00, "custo": 17.00, "descricao": "Molho, muçarela, presunto, ovos, cebola e azeitona."},
        "G": {"preco": 58.00, "custo": 22.00, "descricao": "Molho, muçarela, presunto, ovos, cebola e azeitona."},
        "GG": {"preco": 70.00, "custo": 26.50, "descricao": "Molho, muçarela, presunto, ovos, cebola e azeitona."}
    },
    "quatro queijos": {
        "M": {"preco": 50.00, "custo": 20.00, "descricao": "Molho, muçarela, parmesão, gorgonzola e provolone."},
        "G": {"preco": 65.00, "custo": 25.00, "descricao": "Molho, muçarela, parmesão, gorgonzola e provolone."},
        "GG": {"preco": 78.00, "custo": 30.00, "descricao": "Molho, muçarela, parmesão, gorgonzola e provolone."}
    },
    "frango com catupiry": {
        "M": {"preco": 55.00, "custo": 21.00, "descricao": "Molho, muçarela, frango desfiado e catupiry."},
        "G": {"preco": 72.00, "custo": 28.00, "descricao": "Molho, muçarela, frango desfiado e catupiry."},
        "GG": {"preco": 85.00, "custo": 33.00, "descricao": "Molho, muçarela, frango desfiado e catupiry."}
    },
    "chocolate branco com morango": {
        "M": {"preco": 50.00, "custo": 18.00, "descricao": "Chocolate branco cremoso e morangos frescos."},
        "G": {"preco": 65.00, "custo": 24.00, "descricao": "Chocolate branco cremoso e morangos frescos."},
        "GG": {"preco": 78.00, "custo": 29.00, "descricao": "Chocolate branco cremoso e morangos frescos."}
    },
    "chocolate preto com confetti": {
        "M": {"preco": 48.00, "custo": 17.50, "descricao": "Chocolate meio amargo e granulados coloridos."},
        "G": {"preco": 62.00, "custo": 22.50, "descricao": "Chocolate meio amargo e granulados coloridos."},
        "GG": {"preco": 75.00, "custo": 28.00, "descricao": "Chocolate meio amargo e granulados coloridos."}
    },
    "camarão com catupiry": {
        "M": {"preco": 65.00, "custo": 28.00, "descricao": "Molho, muçarela, camarão salteado e catupiry."},
        "G": {"preco": 85.00, "custo": 35.00, "descricao": "Molho, muçarela, camarão salteado e catupiry."},
        "GG": {"preco": 100.00, "custo": 42.00, "descricao": "Molho, muçarela, camarão salteado e catupiry."}
    }
}

# --- FUNÇÕES DE BANCO DE DADOS ---

def setup_database():
    """Cria o banco de dados SQLite e a tabela 'pedidos' se não existirem."""
    conn = sqlite3.connect('pizzaria.db')
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS pedidos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero_do_dia INTEGER NOT NULL,
            timestamp DATETIME NOT NULL,
            itens_pedido TEXT NOT NULL,
            total_pedido REAL NOT NULL,
            lucro_pedido REAL NOT NULL,
            metodo_pagamento TEXT,
            endereco TEXT,
            observacoes TEXT
        )
    ''')
    conn.commit()
    conn.close()
    print("Banco de dados SQLite pronto.")

def obter_proximo_numero_pedido_dia():
    """Consulta o DB para encontrar o último número de pedido de hoje e retorna o próximo."""
    try:
        conn = sqlite3.connect('pizzaria.db')
        cursor = conn.cursor()
        cursor.execute("SELECT MAX(numero_do_dia) FROM pedidos WHERE DATE(timestamp) = DATE('now', 'localtime')")
        ultimo_numero = cursor.fetchone()[0]
        conn.close()
        return (ultimo_numero or 0) + 1
    except Exception as e:
        print(f"ERRO ao obter número do pedido do dia: {e}")
        return 1

def calcular_lucro_total_dia():
    """Soma a coluna de lucro de todos os pedidos de hoje para obter o total diário."""
    try:
        conn = sqlite3.connect('pizzaria.db')
        cursor = conn.cursor()
        cursor.execute("SELECT SUM(lucro_pedido) FROM pedidos WHERE DATE(timestamp) = DATE('now', 'localtime')")
        lucro_total = cursor.fetchone()[0]
        conn.close()
        return lucro_total or 0.0
    except Exception as e:
        print(f"ERRO ao calcular lucro do dia: {e}")
        return 0.0

def registrar_pedido_sqlite(dados_pedido):
    """Salva um pedido finalizado no banco de dados SQLite."""
    try:
        conn = sqlite3.connect('pizzaria.db')
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO pedidos (numero_do_dia, timestamp, itens_pedido, total_pedido, lucro_pedido, metodo_pagamento, endereco, observacoes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            dados_pedido['numero_do_dia'],
            dados_pedido['timestamp'],
            json.dumps(dados_pedido['itens']),
            dados_pedido['total'],
            dados_pedido['lucro'],
            dados_pedido['pagamento'],
            dados_pedido['endereco'],
            dados_pedido['observacoes']
        ))
        conn.commit()
        conn.close()
        print("-> Pedido registrado com sucesso no SQLite.")
    except Exception as e:
        print(f"ERRO AO REGISTRAR NO SQLITE: {e}")

def registrar_pedido_google_sheets(dados_pedido):
    """Salva e estiliza um pedido na Planilha Google."""
    try:
        gc = gspread.service_account(filename='credentials.json')
        spreadsheet = gc.open("Pedidos da Pizzaria")
        worksheet = spreadsheet.sheet1
        
        all_values = worksheet.get_all_values()
        is_new_sheet = len(all_values) == 0

        if is_new_sheet:
            cabecalho = ["Nº Pedido (Dia)", "Data", "Hora", "Itens", "Total (R$)", "Lucro (R$)", "Pagamento", "Endereço", "Observações"]
            worksheet.append_row(cabecalho)
        
        itens_formatados = ", ".join([f"{item['quantidade']}x {item['sabor']} ({item['tamanho']})" for item in dados_pedido['itens']])
        nova_linha = [
            dados_pedido['numero_do_dia'],
            dados_pedido['timestamp'].strftime("%d/%m/%Y"),
            dados_pedido['timestamp'].strftime("%H:%M:%S"),
            itens_formatados,
            dados_pedido['total'],
            dados_pedido['lucro'],
            dados_pedido['pagamento'],
            dados_pedido['endereco'],
            dados_pedido['observacoes']
        ]
        
        worksheet.append_row(nova_linha)
        print("-> Pedido registrado com sucesso na Planilha Google.")

        nova_linha_numero = len(worksheet.get_all_values())
        
        cor_verde_agua = {"red": 0.85, "green": 0.96, "blue": 0.93}
        borda_solida = {"style": "SOLID", "width": 1, "color": {"red": 0.7, "green": 0.7, "blue": 0.7}}
        
        requests = []
        requests.append({"repeatCell": {"range": {"sheetId": worksheet.id, "startRowIndex": nova_linha_numero - 1, "endRowIndex": nova_linha_numero}, "cell": {"userEnteredFormat": {"backgroundColor": cor_verde_agua, "borders": {"top": borda_solida, "bottom": borda_solida, "left": borda_solida, "right": borda_solida}}}, "fields": "userEnteredFormat(backgroundColor,borders)"}})
        if is_new_sheet:
            requests.append({"repeatCell": {"range": {"sheetId": worksheet.id, "startRowIndex": 0, "endRowIndex": 1}, "cell": {"userEnteredFormat": {"backgroundColor": {"red": 0.49, "green": 0.75, "blue": 0.65}, "textFormat": {"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}}, "borders": {"top": borda_solida, "bottom": borda_solida, "left": borda_solida, "right": borda_solida}}}, "fields": "userEnteredFormat(backgroundColor,textFormat,borders)"}})
            colunas_largura = [{"range": {"sheetId": worksheet.id, "dimension": "COLUMNS", "startIndex": 3, "endIndex": 4}, "properties": {"pixelSize": 350}}, {"range": {"sheetId": worksheet.id, "dimension": "COLUMNS", "startIndex": 4, "endIndex": 6}, "properties": {"pixelSize": 120}}, {"range": {"sheetId": worksheet.id, "dimension": "COLUMNS", "startIndex": 7, "endIndex": 8}, "properties": {"pixelSize": 400}}, {"range": {"sheetId": worksheet.id, "dimension": "COLUMNS", "startIndex": 8, "endIndex": 9}, "properties": {"pixelSize": 250}}]
            for col in colunas_largura:
                requests.append({"updateDimensionProperties": {**col, "fields": "pixelSize"}})
        if requests:
            spreadsheet.batch_update({"requests": requests})
        worksheet.format(f'E{nova_linha_numero}:F{nova_linha_numero}', {"numberFormat": {"type": "CURRENCY", "pattern": "R$ #,##0.00"}})

    except Exception as e:
        print(f"\n!!! ERRO AO INTERAGIR COM GOOGLE SHEETS: {e}\n")

def registrar_pedido_excel(dados_pedido):
    """Salva o pedido em uma planilha Excel diária, localmente."""
    arquivo_excel = f'pedidos_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
    itens_formatados = ", ".join([f"{item['quantidade']}x {item['sabor']} ({item['tamanho']})" for item in dados_pedido['itens']])
    
    nova_linha = [
        dados_pedido['numero_do_dia'],
        dados_pedido['timestamp'].strftime("%H:%M:%S"),
        itens_formatados,
        dados_pedido['total'],
        dados_pedido['lucro'],
        dados_pedido['pagamento'],
        dados_pedido['endereco'],
        dados_pedido['observacoes']
    ]
    
    try:
        if not os.path.exists(arquivo_excel):
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Pedidos"
            cabecalho = ["Nº Pedido (Dia)", "Hora", "Itens", "Total (R$)", "Lucro (R$)", "Pagamento", "Endereço", "Observações"]
            sheet.append(cabecalho)
        else:
            workbook = load_workbook(arquivo_excel)
            if "Pedidos" in workbook.sheetnames:
                sheet = workbook["Pedidos"]
            else:
                sheet = workbook.active

        sheet.append(nova_linha)
        workbook.save(arquivo_excel)
        print("-> Pedido registrado com sucesso no Excel local.")
    except Exception as e:
        print(f"\n!!! ERRO AO REGISTRAR NO EXCEL LOCAL: {e}")
        print(f"AVISO: Verifique se o arquivo '{arquivo_excel}' não está ABERTO em outro programa.\n")


# --- FUNÇÕES DE MENU E SABORES ---
def saudacao_string():
    return (
        "Olá! Bem-vindo à Pizzaria Delícia!\n"
        "Sou seu assistente virtual. Como posso ajudar?\n"
        "1. Ver Cardápio\n"
        "2. Fazer Pedido\n"
        "3. Falar com Atendente\n"
        "4. Sair"
    )

def exibir_menu_opcoes():
    return (
        "1. Ver Cardápio\n"
        "2. Fazer Pedido\n"
        "3. Falar com Atendente\n"
        "4. Sair"
    )

def exibir_cardapio_string():
    linhas_cardapio = ["--- NOSSO CARDÁPIO ---"]
    for sabor, tamanhos in CARDAPIO.items():
        sabor_formatado = sabor.replace("_", " ").title()
        precos = " | ".join([f"{sigla}: R$ {info['preco']:.2f}" for sigla, info in tamanhos.items() if sigla in ["M", "G", "GG"]])
        linhas_cardapio.append(f"- *{sabor_formatado}*: {precos}")
        if "descricao" in tamanhos.get("M", {}):
            linhas_cardapio.append(f"  _{tamanhos['M']['descricao']}_")
    linhas_cardapio.append("----------------------")
    return "\n".join(linhas_cardapio)

def processar_pedido_bot(user_message, session_data):
    """Gerencia o fluxo da conversa de pedido (máquina de estados)."""
    session_data.setdefault('current_order', {})
    session_data.setdefault('order_total', 0.0)

    current_state = session_data.get('state')
    user_input = user_message.lower().strip()
    response_message = ""

    if current_state == 'awaiting_flavor':
        if user_input == 'fim':
            if session_data['current_order']:
                resumo = "Seu pedido foi finalizado com sucesso!\n*Resumo do Pedido:*\n"
                for sabor_full, qtd in session_data['current_order'].items():
                    sabor, tamanho = sabor_full.rsplit(' ', 1)
                    preco = CARDAPIO.get(sabor, {}).get(tamanho.upper(), {}).get("preco", 0)
                    resumo += f"*{qtd}x {sabor.title()} ({tamanho.upper()})* (R$ {preco:.2f} cada)\n"
                resumo += f"\n*Subtotal: R$ {session_data['order_total']:.2f}*\n\n"
                response_message = resumo + "Escolha a forma de pagamento:\n1. Espécie\n2. Cartão\n3. Pix"
                session_data['state'] = 'awaiting_payment_method'
            else:
                response_message = "Nenhum item adicionado. Digite o sabor da pizza ou 'cancelar'."
        elif user_input == 'cancelar':
            chat_history = session_data.get('chat_history', [])
            session_data.clear()
            session_data['chat_history'] = chat_history
            session_data['state'] = 'initial'
            response_message = "Pedido cancelado. Como posso ajudar agora?\n\n" + exibir_menu_opcoes()
        else:
            matched_sabor = next((key for key in CARDAPIO if user_input in key), None)
            if matched_sabor:
                session_data['temp_flavor'] = matched_sabor
                session_data['state'] = 'awaiting_size'
                response_message = f"Certo, *{matched_sabor.title()}*! Qual o tamanho? (M, G, GG)"
            else:
                response_message = "Não encontrei este sabor. Tente novamente ou peça o *'cardápio'*."
    
    elif current_state == 'awaiting_size':
        tamanho_input = user_input.upper()
        sabor_selecionado = session_data.get('temp_flavor')
        if not sabor_selecionado:
            session_data['state'] = 'awaiting_flavor'
            return "Ocorreu um erro. Por favor, escolha o sabor novamente.", session_data
        
        if tamanho_input in CARDAPIO[sabor_selecionado]:
            session_data['temp_size'] = tamanho_input
            session_data['state'] = 'awaiting_item_quantity'
            response_message = f"Ótimo! *{sabor_selecionado.title()} ({tamanho_input})*. Quantas pizzas?"
        else:
            response_message = f"Tamanho inválido para *{sabor_selecionado.title()}*. Escolha M, G ou GG."

    elif current_state == 'awaiting_item_quantity':
        sabor = session_data.get('temp_flavor')
        tamanho = session_data.get('temp_size')
        if not sabor or not tamanho:
            session_data['state'] = 'awaiting_flavor'
            return "Ocorreu um erro. Por favor, comece a adicionar o item novamente.", session_data
        
        try:
            quantidade = int(user_input)
            if quantidade > 0:
                preco = CARDAPIO[sabor][tamanho]["preco"]
                chave_pedido = f"{sabor} {tamanho.lower()}"
                session_data['current_order'][chave_pedido] = session_data['current_order'].get(chave_pedido, 0) + quantidade
                session_data['order_total'] += preco * quantidade
                session_data.pop('temp_flavor', None)
                session_data.pop('temp_size', None)
                session_data['state'] = 'awaiting_flavor'
                response_message = f"*{quantidade}x {sabor.title()} ({tamanho})* adicionada(s).\n\nDigite outro sabor ou *'fim'* para finalizar."
            else:
                response_message = "Por favor, insira um número maior que zero."
        except ValueError:
            response_message = "Entrada inválida. Por favor, digite um número."

# --- FUNÇÕES DE TEXTO PAGAMENTO ---
    elif current_state == 'awaiting_payment_method':
        pagamento = user_input
        if pagamento in ['1', 'espécie', 'especie']:
            session_data['payment_method'] = 'Espécie'
            session_data['state'] = 'awaiting_change_needed'
            response_message = "Pagamento em *Espécie*. Precisará de troco? (Ex: 'Sim, para 50' ou 'Não')"
        elif pagamento in ['2', 'cartão', 'cartao']:
            session_data['payment_method'] = 'Cartão'
            session_data['state'] = 'awaiting_address'
            response_message = "Pagamento com *Cartão*. Por favor, informe seu endereço para entrega."
        elif pagamento in ['3', 'pix']:
            session_data['payment_method'] = 'Pix'
            session_data['state'] = 'awaiting_address'
            pix_key = os.getenv("PIX_KEY", "111.222.333-44 (CPF)")
            response_message = f"Pagamento via *Pix*. Chave: {pix_key}\n\nApós pagar, informe seu endereço."
        else:
            response_message = "Opção inválida. Escolha 1, 2 ou 3."

    elif current_state == 'awaiting_change_needed':
        if "não" in user_input or "nao" in user_input:
            session_data['change_needed'] = "Não precisa de troco."
            session_data['state'] = 'awaiting_address'
            response_message = "Certo. Por favor, informe seu endereço para entrega."
        elif "sim" in user_input:
            try:
                valor_str = user_input.split('para')[-1].strip().replace(',', '.')
                valor_troco = float(valor_str)
                if valor_troco > session_data['order_total']:
                    session_data['change_needed'] = f"Sim, troco para R$ {valor_troco:.2f}"
                    session_data['state'] = 'awaiting_address'
                    response_message = f"Ok, troco para R$ {valor_troco:.2f}. Agora, informe seu endereço."
                else:
                    response_message = "O valor para troco deve ser maior que o total do pedido."
            except (ValueError, IndexError):
                response_message = "Não entendi. Digite 'Sim, para [valor]' ou 'Não'."
        else:
            response_message = "Não entendi. Por favor, digite 'Sim, para [valor]' ou 'Não'."

# --- FUNÇÕES DE TEXTO ENDEREÇO ---
    elif current_state == 'awaiting_address':
        session_data['address'] = user_message
        try:
            custo_total_do_pedido = 0
            itens_para_salvar = []
            for sabor_full, qtd in session_data['current_order'].items():
                sabor_base, tamanho = sabor_full.rsplit(' ', 1)
                custo_item = CARDAPIO.get(sabor_base, {}).get(tamanho.upper(), {}).get("custo", 0)
                custo_total_do_pedido += custo_item * qtd
                itens_para_salvar.append({"sabor": sabor_base.title(), "tamanho": tamanho.upper(), "quantidade": qtd})
            
            lucro_do_pedido = session_data['order_total'] - custo_total_do_pedido
            numero_pedido_dia = obter_proximo_numero_pedido_dia()

            dados_para_registro = {
                "numero_do_dia": numero_pedido_dia, "timestamp": datetime.now(),
                "itens": itens_para_salvar, "total": session_data['order_total'],
                "lucro": lucro_do_pedido, "pagamento": session_data.get('payment_method', 'N/A'),
                "endereco": session_data['address'], "observacoes": session_data.get('change_needed', '')
            }

# --- FUNÇÕES DE REGISTRAR DADOS ---
registrar_pedido_sqlite(dados_para_registro)
            registrar_pedido_google_sheets(dados_para_registro)
            registrar_pedido_excel(dados_para_registro)
# ---------------------------------#

            lucro_total_do_dia = calcular_lucro_total_dia()
            print(f"💰 Lucro total do dia até agora: R$ {lucro_total_do_dia:.2f}")

        except Exception as e:
            print(f"!!! FALHA CRÍTICA AO PROCESSAR DADOS PARA REGISTRO: {e}")

        response_message = f"Endereço: _{session_data['address']}_ confirmado!\n\n*Seu pedido foi enviado para nossa equipe!* Em breve você receberá a confirmação final. *Obrigado por escolher a Pizzaria Delícia!* 🍕"
        if session_data.get('change_needed') and isinstance(session_data.get('change_needed'), str) and "Sim" in session_data.get('change_needed'):
            response_message += f"\n_{session_data['change_needed']}_"
        
        chat_history = session_data.get('chat_history', [])
        session_data.clear()
        session_data['chat_history'] = chat_history
        session_data['state'] = 'initial'
    
    if not response_message:
        response_message = "Desculpe, não entendi. Por favor, escolha uma opção do menu."
        session_data['state'] = 'initial'

    return response_message, session_data

def resumir_conversa_com_ia_terminal(chat_history):
    """Gera um resumo da conversa usando a IA do Gemini."""
    if not genai_model: return "Resumo da IA não disponível."
    try:
        prompt = "Resuma a seguinte conversa para um atendente, focando no pedido do cliente e em qualquer dificuldade encontrada. Seja breve e direto:\n\n" + \
                 "\n".join([f"{msg['role']}: {msg['parts'][0]['text']}" for msg in chat_history])
        response = genai_model.generate_content(prompt)
        return response.text.strip()
    except Exception as e:
        return f"Não foi possível gerar um resumo automático: {e}"

def main_terminal():
    """Função principal que executa o loop do chatbot no terminal."""
    setup_database()
    current_user_session = {'state': 'initial', 'chat_history': []}
    print("Iniciando o Chatbot de Pizzaria no Terminal...")
    print("\n" + saudacao_string())

    while True:
        user_input = input("Você: ").strip()
        current_user_session['chat_history'].append({"role": "user", "parts": [{"text": user_input}]})
        bot_response = ""
        user_input_normalized = user_input.lower()

        if current_user_session.get('state') == 'initial':
            if "olá" in user_input_normalized or "oi" in user_input_normalized:
                bot_response = saudacao_string()
            elif user_input_normalized == '1':
                bot_response = f"🍕 *Nosso Cardápio:* 🍕\n\n{exibir_cardapio_string()}\n\nPara pedir, digite *'2'*."
            elif user_input_normalized == '2':
                sabores = ", ".join([f"*{s.replace('_',' ').title()}*" for s in CARDAPIO.keys()])
                bot_response = f"✨ *Vamos montar seu pedido!* ✨\n\n*Sabores:* {sabores}\n\nQual sabor você gostaria?"
                current_user_session['state'] = 'awaiting_flavor'
            elif user_input_normalized == '3':
                bot_response = "Ok! Um atendente foi notificado com o resumo da nossa conversa.\n\n"
                bot_response += f"--- RESUMO ---\n{resumir_conversa_com_ia_terminal(current_user_session['chat_history'])}"
            elif user_input_normalized == '4':
                bot_response = "Obrigado por usar nosso serviço! Volte sempre! 👋"
                print(f"Bot: {bot_response}")
                break
            else:
                bot_response = "Opção inválida. Por favor, escolha uma das opções (1-4)."
        
        elif current_user_session.get('state', '').startswith('awaiting_'):
            bot_response, current_user_session = processar_pedido_bot(user_input, current_user_session)
            
        current_user_session['chat_history'].append({"role": "model", "parts": [{"text": bot_response}]})
        print(f"Bot: {bot_response}")

if __name__ == "__main__":
    main_terminal()